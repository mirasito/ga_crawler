"""Shared pytest fixtures for Phase 3 (Goldapple Crawl) tests.

Loads spike sample-payloads as in-memory fixtures. Provides Phase 2 contract
mocks (BrandAlias, Normalizer, SnapshotWriter, RunWriter) for all unit and
mocked-integration tests. Tmp Camoufox profile dir factory for fetcher tests.

Fixture path conventions:
- HTML fixtures: tests/fixtures/goldapple/*.html (verbatim spike copies)
- Mock builders: returned via fixture functions, scope='function' (fresh per test)
- Tmp profile dirs: scope='function', auto-cleanup via tmp_path

Per 03-PATTERNS.md "Test files" table.
"""

from __future__ import annotations

import json
import re
import warnings
from datetime import datetime, timedelta, timezone
from pathlib import Path
from unittest.mock import MagicMock

import pytest

FIXTURES_DIR = Path(__file__).parent / "fixtures" / "goldapple"

# ============================================================================
# Phase 9 Wave 0: PII canary + fixture-loader integration (D-907)
# ============================================================================

_PII_PATTERNS: tuple[re.Pattern[str], ...] = (
    re.compile(r"cf_clearance\s*=", re.IGNORECASE),
    re.compile(r"\bbot\d{9,10}:[A-Za-z0-9_\-]{30,}\b"),
    re.compile(r"\bAuthorization:\s*Bearer\b", re.IGNORECASE),
    # hc-ping paths embed UUID-shaped healthcheck IDs — PII-adjacent (operator token).
    # Standalone UUID v4 is intentionally EXCLUDED: goldapple HTML legitimately embeds
    # buildId values in UUID v4 format (e.g. Nuxt buildId in __NEXT_DATA__). The
    # hc-ping-specific pattern below covers the actual hc-ping threat.
    re.compile(r"hc-ping\.com/[0-9a-f-]{32,36}", re.IGNORECASE),
)
_MAX_FIXTURE_BYTES = 50 * 1024 * 1024  # 50 MB per file
_MAX_AGGREGATE_BYTES = 200 * 1024 * 1024  # 200 MB aggregate


def _assert_fixture_clean(path: Path) -> None:
    """D-907 fail-fast canary: PII pattern + 50 MB size budget per fixture.

    Fails with path + matched-pattern (NOT matched-text) so secrets don't
    leak into pytest stdout if accidentally captured. T-09-PII + T-09-SIZE.
    """
    import pytest as _pytest  # local import: avoid top-of-file shuffle

    size = path.stat().st_size
    if size > _MAX_FIXTURE_BYTES:
        _pytest.fail(
            f"Fixture {path.name} is {size} bytes (>{_MAX_FIXTURE_BYTES}); "
            f"refuse to load (50 MB per-file budget)."
        )
    content = path.read_text(encoding="utf-8", errors="replace")
    for pat in _PII_PATTERNS:
        m = pat.search(content)
        if m:
            _pytest.fail(
                f"PII pattern matched in {path}: pattern={pat.pattern!r}. "
                f"Scrub fixture and recapture (do NOT commit secret content)."
            )


def _check_fixture_age(sidecar_path: Path) -> None:
    """Stale-fixture UserWarning at 30 days per D-906."""
    if not sidecar_path.exists():
        return
    import json as _json
    try:
        meta = _json.loads(sidecar_path.read_text(encoding="utf-8"))
        captured = datetime.fromisoformat(meta["date"])
        if captured.tzinfo is None:
            captured = captured.replace(tzinfo=timezone.utc)
        age = datetime.now(timezone.utc) - captured
        if age > timedelta(days=30):
            warnings.warn(
                f"Live fixture {sidecar_path.stem} is {age.days}d old (>30d). "
                f"Consider `pytest -m live --refresh-live --snapshot-update`.",
                UserWarning,
                stacklevel=2,
            )
    except (KeyError, ValueError, OSError):
        # Sidecar absent or malformed — silent. Standalone canary covers shape.
        pass


# ---- HTML / sitemap fixtures ----

@pytest.fixture(scope="session")
def goldapple_pdp_html() -> str:
    """Real Givenchy PDP captured during spike 01-08 (~200 KB).

    Source: .planning/spikes/01-goldapple/sample-payloads/_debug-product-page.html
    Used by parser tests to assert: top-level offer extraction, priceType filtering
    (StrikethroughPrice / ListPrice / Gold Card discrimination), brand microdata
    walk, availability schema.org URL → enum.
    """
    return (FIXTURES_DIR / "_debug-product-page.html").read_text(encoding="utf-8")


@pytest.fixture(scope="session")
def gate_shell_html() -> str:
    """GroupIB challenge shell sample (~18 KB).

    Source: .planning/spikes/01-goldapple/sample-payloads/goldapple-product-html-1.html
    Used by gate detection tests: title contains "checking device" + size < 30 KB.
    """
    return (FIXTURES_DIR / "gate-shell.html").read_text(encoding="utf-8")


@pytest.fixture(scope="session")
def stale_sku_html() -> str:
    """De-listed SKU pattern: 200 OK + ~9.5 KB + no microdata + title contains 'Loading'.

    Source: synthesized at Wave 0 to mimic spike result row 0 (per
    03-PATTERNS.md §"Test fixtures" — stale-sku-9.5kb.html "must be re-fetched
    or synthesized").
    """
    return (FIXTURES_DIR / "stale-sku-9.5kb.html").read_text(encoding="utf-8")


@pytest.fixture(scope="session")
def sitemap_xml() -> str:
    """Goldapple sub-sitemap excerpt (50 product URLs).

    Source: .planning/spikes/01-goldapple/sample-payloads/goldapple-sitemap-1-excerpt.xml
    Used by sitemap parser unit tests (regex extraction of <loc>...</loc>).
    """
    return (FIXTURES_DIR / "sitemap-1-excerpt.xml").read_text(encoding="utf-8")


@pytest.fixture(scope="session")
def tier2_results_json() -> dict:
    """Spike 01-08 empirical 100-fetch baseline (99/100 success, 0% gate-shell).

    Source: .planning/spikes/01-goldapple/sample-payloads/tier2-camoufox-kz-results.json
    Used by stale-SKU detection tests: row 0 = '7681000002-...' is the canonical
    stale-SKU case (status=200, html_size~9500, no microdata).
    """
    return json.loads((FIXTURES_DIR / "tier2-camoufox-kz-results.json").read_text(encoding="utf-8"))


@pytest.fixture(scope="session")
def jsonld_blocks_anti_fixture() -> list:
    """Proof goldapple emits ONLY OfferShippingDetails JSON-LD (no Product schema).

    Source: .planning/spikes/01-goldapple/sample-payloads/_debug-jsonld-blocks.json
    Used by parser tests as anti-fixture: assert parser does NOT use JSON-LD path
    for goldapple (D-14 microdata-not-JSON-LD; PARSE-02 inverted for goldapple).
    """
    return json.loads((FIXTURES_DIR / "_debug-jsonld-blocks.json").read_text(encoding="utf-8"))


# ---- Phase 8 v1.1 live fixtures (Plan 08-01 W0 spike) ----

@pytest.fixture(scope="session")
def goldapple_pdp_html_live_stereotype() -> str:
    """STEREOTYPE / SAĜO live PDP captured 2026-05-13 (Bug #1+#2 evidence).

    h1 .brand = "Stereotype", h1 .name = "SAĜO" — Title-case + UPPERCASE.
    Volume block [75] ОБЪЁМ / МЛ present.
    Source: Plan 08-01 W0 spike (.planning/spikes/v1.1-brand-name-shapes/
    pdp-16-niche-stereotype-sago.html), Camoufox 0.4.11 fetch.
    """
    path = FIXTURES_DIR / "_live-2026-05-13-stereotype.html"
    _assert_fixture_clean(path)                    # D-907 enforcement point #1
    _check_fixture_age(path.with_suffix(".json"))  # D-906 stale-fixture warning
    return path.read_text(encoding="utf-8")


@pytest.fixture(scope="session")
def goldapple_pdp_html_live_armani() -> str:
    """Armani Code live PDP captured 2026-05-13 (Bug #2 evidence).

    h1 .brand = "Armani", h1 .name = "armani code" — brand string is a
    substring of name (data redundancy from upstream goldapple catalog).
    Source: Plan 08-01 W0 spike (pdp-07-mass-armani-code.html).
    """
    path = FIXTURES_DIR / "_live-2026-05-13-armani-code.html"
    _assert_fixture_clean(path)                    # D-907 enforcement point #1
    _check_fixture_age(path.with_suffix(".json"))  # D-906 stale-fixture warning
    return path.read_text(encoding="utf-8")


# ---- Phase 2 contract mocks ----

@pytest.fixture
def mock_brand_alias() -> MagicMock:
    """Mock BrandAliasProtocol. Default lookup returns the brand_norm itself as
    the only alias; tests override .lookup.return_value or .side_effect as needed.
    """
    mock = MagicMock()
    mock.lookup.side_effect = lambda b: [b]
    return mock


@pytest.fixture
def mock_normalizer() -> MagicMock:
    """Mock NormalizerProtocol. Default brand/name = lowercase + strip; volume = None."""
    mock = MagicMock()
    mock.brand.side_effect = lambda raw: raw.lower().strip()
    mock.name.side_effect = lambda raw: raw.lower().strip()
    mock.volume.return_value = None
    return mock


@pytest.fixture
def mock_snapshot_writer() -> MagicMock:
    """Mock SnapshotWriterProtocol. .append returns len(products); records calls
    via mock.append.call_args_list for assertion."""
    mock = MagicMock()
    mock.append.side_effect = lambda run_id, retailer, products: len(products)
    return mock


@pytest.fixture
def mock_run_writer() -> MagicMock:
    """Mock RunWriterProtocol. patch_stats accumulates into _stats dict;
    get_stats returns it; fail records reason."""
    mock = MagicMock()
    mock._stats = {}

    def _patch(run_id, delta):
        mock._stats.update(delta)

    def _get(run_id):
        return dict(mock._stats)

    mock.patch_stats.side_effect = _patch
    mock.get_stats.side_effect = _get
    return mock


# ---- Tmp profile dir for fetcher tests ----

@pytest.fixture
def tmp_camoufox_profile_dir(tmp_path: Path) -> Path:
    """Per-test Camoufox tmp profile dir under pytest's tmp_path.
    Auto-cleanup by pytest at end of test. Mirrors D-311 fresh-profile-per-run pattern."""
    p = tmp_path / "camoufox-profile"
    p.mkdir()
    return p


# ===== Phase 2 fixtures (D-222) =====
# Added Wave 0 of Phase 2. The 11 Phase 3 fixtures above remain untouched.
# Source: 02-CONTEXT.md D-222 + 02-PATTERNS.md §"Pattern: conftest fixture extension".

import yaml  # noqa: E402  -- import-after-block intentional for section grouping
from sqlmodel import Session, SQLModel, create_engine  # noqa: E402

VILED_FIXTURES_DIR = Path(__file__).parent / "fixtures" / "viled"
NORMALIZE_FIXTURES_DIR = Path(__file__).parent / "fixtures" / "normalize"


@pytest.fixture(scope="session")
def viled_pdp_html() -> str:
    """Canonical viled PDP HTML pinned Wave 0 (item/407682, Alice+Olivia "Кружевное боди").

    Source: tests/fixtures/viled/viled-pdp-407682.html captured 2026-05-07 via
    curl_cffi.requests.get(impersonate="chrome"); see 02-WAVE0-PROBE.md.

    Drives PARSE-01..04 happy-path tests for the viled __NEXT_DATA__ parser.
    """
    return (VILED_FIXTURES_DIR / "viled-pdp-407682.html").read_text(encoding="utf-8")


@pytest.fixture(scope="session")
def viled_pdp_discounted_html() -> str:
    """Discounted viled PDP HTML (item/367251, Frederic Malle perfume).

    attributes[0].price=356745 < attributes[0].realPrice=419700, enableDiscount=True.
    Drives PARSE-03 was_price assertion.
    """
    return (VILED_FIXTURES_DIR / "viled-pdp-discounted.html").read_text(encoding="utf-8")


@pytest.fixture(scope="session")
def viled_pdp_multipack_html() -> str:
    """Multipack viled PDP HTML (item/398309). Drives NORM-04 multipack-flag tests."""
    return (VILED_FIXTURES_DIR / "viled-pdp-multipack.html").read_text(encoding="utf-8")


@pytest.fixture(scope="session")
def viled_pdp_html_live_contre_jour() -> str:
    """Frederic Malle / Contre-Jour live PDP captured 2026-05-13 (Bug #3 evidence).

    viled.kz/item/408872. Drives PARSE-FIX-03 legitimate-None case per D-814:
    `Размер` attribute may be absent → volume_norm stays NULL after fallback.
    Source: Plan 08-01 W0 spike (.planning/spikes/v1.1-brand-name-shapes/
    viled-contre-jour-408872.html), curl_cffi fetch.
    """
    path = VILED_FIXTURES_DIR / "_live-2026-05-13-contre-jour.html"
    _assert_fixture_clean(path)                    # D-907 enforcement point #1
    _check_fixture_age(path.with_suffix(".json"))  # D-906 stale-fixture warning
    return path.read_text(encoding="utf-8")


@pytest.fixture(scope="session")
def viled_catalog_html() -> str:
    """Canonical viled catalog page-1 HTML (men/catalog/1310, ~238 KB, 60 items).

    Source: tests/fixtures/viled/viled-catalog-men-1310-page1.html.
    Drives CRAWL-01 catalog enumeration tests; pageProps.items.{content,total,
    totalPages,pageSize,pageNumber} per 02-WAVE0-PROBE.md A4.
    """
    return (VILED_FIXTURES_DIR / "viled-catalog-men-1310-page1.html").read_text(encoding="utf-8")


@pytest.fixture
def brand_alias_yaml_fixture(tmp_path: Path) -> Path:
    """Materialize tests/fixtures/viled/brand-aliases-fixture.yaml into tmp_path.

    Returns the tmp file path so YamlBrandAlias tests can point at a writable copy
    (allows mutation in tests without polluting the source fixture).
    Schema per 02-CONTEXT.md D-205.
    """
    src = (VILED_FIXTURES_DIR / "brand-aliases-fixture.yaml").read_text(encoding="utf-8")
    dst = tmp_path / "brand-aliases.yaml"
    dst.write_text(src, encoding="utf-8")
    return dst


@pytest.fixture
def in_memory_sqlite_session():
    """Yield a Session bound to an in-memory SQLite engine with foreign-keys
    PRAGMA applied. SQLModel tables are created if Wave 1's storage module is
    importable; otherwise a bare session yields (Wave 0 stubs that need this
    fixture must skip until Wave 1 lands).

    Source: 02-RESEARCH.md §Pattern 3 + §"Initializing the WAL session".
    """
    from sqlalchemy import event

    engine = create_engine("sqlite:///:memory:", echo=False)

    @event.listens_for(engine, "connect")
    def _set_pragma(dbapi_conn, _):
        cur = dbapi_conn.cursor()
        # WAL is moot for :memory: but PRAGMA call is harmless and matches prod.
        cur.execute("PRAGMA foreign_keys=ON")
        cur.close()

    # Tables are created lazily by Wave 1; Wave 0 stub yields a bare session.
    try:
        from ga_crawler.storage.sqlite import SQLModel as _SQL  # type: ignore[import-not-found]

        _SQL.metadata.create_all(engine)
    except ImportError:
        SQLModel.metadata.create_all(engine)  # empty metadata, no-op until Wave 1
    with Session(engine) as session:
        yield session


@pytest.fixture
def volume_corpus_cases() -> list[dict]:
    """Load tests/fixtures/normalize/volume-corpus.yaml.

    Returns list[dict] with keys input/expected_volume/expected_multipack.
    Source: 02-RESEARCH.md Open Q5 + Pattern 6.
    """
    raw = (NORMALIZE_FIXTURES_DIR / "volume-corpus.yaml").read_text(encoding="utf-8")
    return yaml.safe_load(raw)["cases"]


@pytest.fixture
def brand_corpus_cases() -> list[dict]:
    """Load tests/fixtures/normalize/brand-corpus.yaml.

    Returns list[dict] with keys raw/expected_brand_norm/alias_present.
    """
    raw = (NORMALIZE_FIXTURES_DIR / "brand-corpus.yaml").read_text(encoding="utf-8")
    return yaml.safe_load(raw)["cases"]


# =====================================================================
# Phase 5 reporter fixtures — appended per D-222 append-only pattern.
# Closest analog: synthetic data planting in tests/integration/test_matcher_run.py
# (Phase 4). Used by Plans 05-02 / 05-03 / 05-04 / 05-05.
# Source: 05-VALIDATION.md Wave 0 Requirements + 05-PATTERNS.md
# "tests/conftest.py extension" section.
# =====================================================================

from datetime import datetime, timezone  # noqa: E402  -- import-after-block intentional


@pytest.fixture
def tmp_reports_dir(tmp_path):
    """tmp_path-based output_dir for Phase 5 archive tests.

    Returns a Path to `tmp_path/reports/` (already mkdir'd). Mirror of
    how Plan 04 tests use tmp_path engine — keeps each test isolated.
    """
    p = tmp_path / "reports"
    p.mkdir(parents=True, exist_ok=True)
    return p


@pytest.fixture
def openpyxl_workbook_reader():
    """Returns a callable: open xlsx (bytes or Path) → openpyxl.Workbook.

    Used by excel_builder/reporter_run tests to assert sheet names,
    freeze_panes coord, autofilter range, and conditional_formatting rules
    on the xlsx produced by xlsxwriter.

    Source: 05-RESEARCH.md Pitfall 3 — assert *behavioral* structure
    (sheet existence + freeze coord + cf rule type) NOT exact hex colors.
    """
    import io as _io

    from openpyxl import load_workbook

    def _open(src):
        if isinstance(src, (bytes, bytearray)):
            return load_workbook(_io.BytesIO(src), read_only=False)
        return load_workbook(str(src), read_only=False)

    return _open


@pytest.fixture
def synthetic_report_run(tmp_path):
    """In-memory SQLite engine + 1 Run + paired viled/goldapple snapshots + matches + promos.

    Returns: (engine, run_writer, run_id, repo_root)

    Populated state — week-1 baseline test fixture for Plans 05-02/03/04/05:
      - 1 Run row (status='success', started_at = 2026-05-10 14:00 UTC → ISO 2026-W19
        in Asia/Almaty)
      - 3 viled snapshots (all matched)
      - 8 goldapple snapshots (3 matched + 3 gap-only + 2 promo gap-only)
      - 3 Match rows with KNOWN price_delta_pct values for top-3 assertion:
          * (givenchy, eau de parfum, 50ml): delta_pct = +15.50  (goldapple more expensive)
          * (creed, aventus, 100ml): delta_pct = -22.30 (viled more expensive)
          * (dior, sauvage, 100ml): delta_pct = +5.00 (goldapple slightly more)
        Top-3 sort by ABS(delta_pct) DESC = creed > givenchy > dior.

    Source: 05-VALIDATION.md Wave 0 Requirements + 05-PATTERNS.md
    "tests/conftest.py extension" section.
    """
    from sqlalchemy import text as _text

    from ga_crawler.storage.sqlite import (
        SqliteRunWriter,
        SqliteSnapshotWriter,
        init_db,
        make_engine,
    )

    db_path = tmp_path / "reporter.db"
    init_db(db_path)
    engine = make_engine(db_path)
    run_writer = SqliteRunWriter(engine)

    # Create run, then UPDATE started_at to deterministic value → ISO 2026-W19 in Asia/Almaty.
    # (SqliteRunWriter.create() does not accept started_at; default uses now(UTC).)
    run_id = run_writer.create()
    started_at = datetime(2026, 5, 10, 14, 0, 0, tzinfo=timezone.utc)
    with engine.begin() as conn:
        conn.execute(
            _text("UPDATE runs SET started_at = :sa WHERE run_id = :rid"),
            {"sa": started_at, "rid": run_id},
        )

    # Plant viled snapshots (3 matched)
    viled_rows = [
        dict(
            sku_id="v-givenchy-edp-50",
            url="https://viled.kz/p/givenchy-edp-50",
            name="Givenchy Eau de Parfum 50ml",
            brand="Givenchy",
            volume_raw="50 мл",
            current_price=50000,
            was_price=None,
            currency="KZT",
            stock_state="IN_STOCK",
            brand_norm="givenchy",
            name_norm="eau de parfum",
            volume_norm="(50, ml, 1)",
            multipack_flag=False,
            scraped_at=datetime(2026, 5, 10, 14, 5, 0, tzinfo=timezone.utc),
        ),
        dict(
            sku_id="v-creed-aventus-100",
            url="https://viled.kz/p/creed-aventus-100",
            name="Creed Aventus 100ml",
            brand="Creed",
            volume_raw="100 мл",
            current_price=180000,
            was_price=None,
            currency="KZT",
            stock_state="IN_STOCK",
            brand_norm="creed",
            name_norm="aventus",
            volume_norm="(100, ml, 1)",
            multipack_flag=False,
            scraped_at=datetime(2026, 5, 10, 14, 6, 0, tzinfo=timezone.utc),
        ),
        dict(
            sku_id="v-dior-sauvage-100",
            url="https://viled.kz/p/dior-sauvage-100",
            name="Dior Sauvage 100ml",
            brand="Dior",
            volume_raw="100 мл",
            current_price=60000,
            was_price=None,
            currency="KZT",
            stock_state="IN_STOCK",
            brand_norm="dior",
            name_norm="sauvage",
            volume_norm="(100, ml, 1)",
            multipack_flag=False,
            scraped_at=datetime(2026, 5, 10, 14, 7, 0, tzinfo=timezone.utc),
        ),
    ]

    # Plant goldapple snapshots: 3 matched + 3 gap-only + 2 promo gap-only (was_price > current)
    goldapple_rows = [
        # 3 matched (same strict key as viled)
        dict(
            sku_id="g-givenchy-edp-50",
            url="https://goldapple.kz/p/givenchy-edp-50",
            name="Givenchy Eau de Parfum 50ml",
            brand="Givenchy",
            volume_raw="50 мл",
            current_price=57750,  # +15.5% vs viled 50000
            was_price=None,
            currency="KZT",
            stock_state="IN_STOCK",
            brand_norm="givenchy",
            name_norm="eau de parfum",
            volume_norm="(50, ml, 1)",
            multipack_flag=False,
            scraped_at=datetime(2026, 5, 10, 14, 10, 0, tzinfo=timezone.utc),
        ),
        dict(
            sku_id="g-creed-aventus-100",
            url="https://goldapple.kz/p/creed-aventus-100",
            name="Creed Aventus 100ml",
            brand="Creed",
            volume_raw="100 мл",
            current_price=139860,  # -22.3% vs viled 180000
            was_price=None,
            currency="KZT",
            stock_state="IN_STOCK",
            brand_norm="creed",
            name_norm="aventus",
            volume_norm="(100, ml, 1)",
            multipack_flag=False,
            scraped_at=datetime(2026, 5, 10, 14, 11, 0, tzinfo=timezone.utc),
        ),
        dict(
            sku_id="g-dior-sauvage-100",
            url="https://goldapple.kz/p/dior-sauvage-100",
            name="Dior Sauvage 100ml",
            brand="Dior",
            volume_raw="100 мл",
            current_price=63000,  # +5% vs viled 60000
            was_price=None,
            currency="KZT",
            stock_state="IN_STOCK",
            brand_norm="dior",
            name_norm="sauvage",
            volume_norm="(100, ml, 1)",
            multipack_flag=False,
            scraped_at=datetime(2026, 5, 10, 14, 12, 0, tzinfo=timezone.utc),
        ),
        # 3 gap-only (no viled counterpart)
        dict(
            sku_id="g-chanel-no5-50",
            url="https://goldapple.kz/p/chanel-no5-50",
            name="Chanel No 5 50ml",
            brand="Chanel",
            volume_raw="50 мл",
            current_price=70000,
            was_price=None,
            currency="KZT",
            stock_state="IN_STOCK",
            brand_norm="chanel",
            name_norm="no 5",
            volume_norm="(50, ml, 1)",
            multipack_flag=False,
            scraped_at=datetime(2026, 5, 10, 14, 13, 0, tzinfo=timezone.utc),
        ),
        dict(
            sku_id="g-armani-acqua-50",
            url="https://goldapple.kz/p/armani-acqua-50",
            name="Armani Acqua di Gio 50ml",
            brand="Armani",
            volume_raw="50 мл",
            current_price=45000,
            was_price=None,
            currency="KZT",
            stock_state="IN_STOCK",
            brand_norm="armani",
            name_norm="acqua di gio",
            volume_norm="(50, ml, 1)",
            multipack_flag=False,
            scraped_at=datetime(2026, 5, 10, 14, 14, 0, tzinfo=timezone.utc),
        ),
        dict(
            sku_id="g-ysl-libre-50",
            url="https://goldapple.kz/p/ysl-libre-50",
            name="YSL Libre 50ml",
            brand="YSL",
            volume_raw="50 мл",
            current_price=55000,
            was_price=None,
            currency="KZT",
            stock_state="IN_STOCK",
            brand_norm="ysl",
            name_norm="libre",
            volume_norm="(50, ml, 1)",
            multipack_flag=False,
            scraped_at=datetime(2026, 5, 10, 14, 15, 0, tzinfo=timezone.utc),
        ),
        # 2 promos (was_price > current_price) — gap-only too, so they show on gap + promo sheets
        dict(
            sku_id="g-tom-ford-noir-50",
            url="https://goldapple.kz/p/tom-ford-noir-50",
            name="Tom Ford Noir 50ml",
            brand="Tom Ford",
            volume_raw="50 мл",
            current_price=80000,
            was_price=100000,  # 20% promo
            currency="KZT",
            stock_state="IN_STOCK",
            brand_norm="tom ford",
            name_norm="noir",
            volume_norm="(50, ml, 1)",
            multipack_flag=False,
            scraped_at=datetime(2026, 5, 10, 14, 16, 0, tzinfo=timezone.utc),
        ),
        dict(
            sku_id="g-givenchy-irresistible-30",
            url="https://goldapple.kz/p/givenchy-irresistible-30",
            name="Givenchy Irresistible 30ml",
            brand="Givenchy",
            volume_raw="30 мл",
            current_price=25000,
            was_price=30000,  # ~16.67% promo
            currency="KZT",
            stock_state="IN_STOCK",
            brand_norm="givenchy",
            name_norm="irresistible",
            volume_norm="(30, ml, 1)",
            multipack_flag=False,
            scraped_at=datetime(2026, 5, 10, 14, 17, 0, tzinfo=timezone.utc),
        ),
    ]

    snap_writer = SqliteSnapshotWriter(engine, batch_size=20)
    snap_writer.append(run_id, "viled", viled_rows)
    snap_writer.append(run_id, "goldapple", goldapple_rows)

    # Plant 3 matches directly (skip running real matcher to keep fixture deterministic).
    # Schema: matches(run_id, viled_sku, goldapple_sku, brand_norm, name_norm,
    #  volume_norm, viled_price, goldapple_price, viled_was_price,
    #  goldapple_was_price, price_delta, price_delta_pct, matched_at)
    matches = [
        (run_id, "v-givenchy-edp-50", "g-givenchy-edp-50", "givenchy", "eau de parfum",
         "(50, ml, 1)", 50000, 57750, None, None, 7750, 15.50, "2026-05-10T14:20:00Z"),
        (run_id, "v-creed-aventus-100", "g-creed-aventus-100", "creed", "aventus",
         "(100, ml, 1)", 180000, 139860, None, None, -40140, -22.30, "2026-05-10T14:20:00Z"),
        (run_id, "v-dior-sauvage-100", "g-dior-sauvage-100", "dior", "sauvage",
         "(100, ml, 1)", 60000, 63000, None, None, 3000, 5.00, "2026-05-10T14:20:00Z"),
    ]
    with engine.begin() as conn:
        conn.execute(
            _text(
                "INSERT INTO matches (run_id, viled_sku, goldapple_sku, brand_norm, "
                "name_norm, volume_norm, viled_price, goldapple_price, viled_was_price, "
                "goldapple_was_price, price_delta, price_delta_pct, matched_at) "
                "VALUES (:rid,:vs,:gs,:bn,:nn,:vn,:vp,:gp,:vw,:gw,:pd,:pp,:mat)"
            ),
            [
                dict(rid=m[0], vs=m[1], gs=m[2], bn=m[3], nn=m[4], vn=m[5],
                     vp=m[6], gp=m[7], vw=m[8], gw=m[9], pd=m[10], pp=m[11], mat=m[12])
                for m in matches
            ],
        )

    # Finalize run as success (D-507 status-gate requirement).
    run_writer.finalize(run_id, "success")

    # Pre-populate runs.stats with upstream namespaces reporter will read.
    # match.count = 3, match.rate = 60.0 (Phase 4 D-405 formula frozen; for this
    # fixture we set the rate directly so reporter tests have a known value).
    run_writer.patch_stats(
        run_id,
        {
            "viled.fetch_count": 3,
            "goldapple.fetch_count": 8,
            "match.count": 3,
            "match.rate": 60.0,
            "match.denominator": 5,
        },
    )

    return engine, run_writer, run_id, tmp_path


# ============================================================================
# Plan 06-01 (Wave 0) — Phase 6 delivery fixtures.
# ============================================================================


@pytest.fixture
def synthetic_delivered_run(synthetic_report_run):
    """Phase 6 Pitfall F superset of synthetic_report_run.

    Populates report.* (xlsx_path + summary_text + size_guard_passed=True)
    on top of the report.* keys planted by synthetic_report_run. Also
    plants a fake xlsx file on disk so FSInputFile would succeed.
    Returns: (engine, run_writer, run_id, repo_root).
    """
    engine, run_writer, run_id, repo_root = synthetic_report_run
    reports_dir = repo_root / "reports"
    reports_dir.mkdir(exist_ok=True)
    fake_xlsx = reports_dir / "2026-W19.xlsx"
    fake_xlsx.write_bytes(b"PK\x03\x04fake-xlsx-content-for-tests")
    run_writer.patch_stats(run_id, {
        "report.xlsx_path": "reports/2026-W19.xlsx",
        "report.xlsx_size_bytes": fake_xlsx.stat().st_size,
        "report.summary_text": "\U0001f4ca Weekly report 2026-W19\nmatch_count: 3 (60.0%)",
        "report.sheet_row_counts": {"summary": 1, "per_sku_deltas": 3},
        "report.skipped_reason": "",
        "report.size_guard_passed": True,
        "report.generated_at": "2026-05-10T14:30:00+00:00",
    })
    return engine, run_writer, run_id, repo_root


@pytest.fixture
def mock_aiogram_bot(mocker):
    """Mock aiogram Bot — send_message/send_document return Message with stub id.

    Returns a MagicMock with __aenter__/__aexit__ as AsyncMock so it works
    as an async context manager. send_message returns message_id=10001;
    send_document returns message_id=10002.
    """
    from unittest.mock import AsyncMock, MagicMock
    bot = MagicMock()
    bot.__aenter__ = AsyncMock(return_value=bot)
    bot.__aexit__ = AsyncMock(return_value=None)
    bot.send_message = AsyncMock(return_value=MagicMock(message_id=10001))
    bot.send_document = AsyncMock(return_value=MagicMock(message_id=10002))
    bot.session = MagicMock()
    bot.session.close = AsyncMock(return_value=None)
    return bot


@pytest.fixture
def mock_tg_env(monkeypatch):
    """Set TG_* env vars for delivery tests.

    Per RESEARCH caveat #4: tests use monkeypatch.setenv, NOT .env file
    loading. Yields a dict of the planted values for easy assertions.
    """
    monkeypatch.setenv("TG_BOT_TOKEN", "test-token-12345")
    monkeypatch.setenv("TG_BUSINESS_CHAT_ID", "-100000001")
    monkeypatch.setenv("TG_OPS_CHAT_ID", "-100000002")
    yield {
        "bot_token": "test-token-12345",
        "business_chat_id": "-100000001",
        "ops_chat_id": "-100000002",
    }


# ============================================================================
# Phase 9 Wave 0 — syrupy + custom CLI flag wiring (Task 4, D-906)
# ============================================================================

from tests._snapshot_extension import HTMLSnapshotExtension  # noqa: E402


def pytest_addoption(parser):
    """D-906 two-mode harness: --refresh-live opts into live re-fetch.

    Verbatim adaptation of Context7-verified --runslow pattern from pytest 8
    docs §"Control test skipping with custom CLI options".
    """
    parser.addoption(
        "--refresh-live",
        action="store_true",
        default=False,
        help=(
            "Re-fetch live URLs via Camoufox/curl_cffi instead of replaying "
            "cassettes. Operator-only; combine with --snapshot-update to "
            "regenerate fixtures."
        ),
    )


@pytest.fixture
def refresh_live(request) -> bool:
    return bool(request.config.getoption("--refresh-live"))


@pytest.fixture
def html_snapshot(snapshot):
    """syrupy snapshot configured to write/read .html files via
    HTMLSnapshotExtension. RESEARCH §3.1."""
    return snapshot.with_defaults(extension_class=HTMLSnapshotExtension)
